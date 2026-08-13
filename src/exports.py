"""Excel export of the full SIOP readout with professional formatting."""
from __future__ import annotations

import io
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import BaselineResult, Recommendation, SimulationResult
from .utils import month_labels, quarter_labels

HEADER_FILL = PatternFill("solid", fgColor="1F3B5C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3B5C")
NOTE_FONT = Font(italic=True, size=10, color="666666")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = "#,##0.0,,\"M\""      # dollars shown in $M
PCT_FMT = "0.0%"
NUM_FMT = "#,##0.0"


def _style_sheet(ws: Worksheet, df: pd.DataFrame, start_row: int = 3,
                 money_cols: tuple[str, ...] = (), pct_cols: tuple[str, ...] = (),
                 heat_col: str | None = None) -> None:
    """Header styling, borders, number formats, filters, frozen panes, widths."""
    n_rows, n_cols = df.shape
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        q = df[col].astype(str).str.len().quantile(0.9)
        width = max(14, min(42, int(q if np.isfinite(q) else 14) + 4))
        ws.column_dimensions[get_column_letter(j)].width = width
        fmt = MONEY_FMT if col in money_cols else PCT_FMT if col in pct_cols else None
        for i in range(start_row + 1, start_row + 1 + n_rows):
            c = ws.cell(row=i, column=j)
            c.border = BORDER
            if fmt and isinstance(c.value, (int, float)):
                c.number_format = fmt
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = (f"A{start_row}:"
                          f"{get_column_letter(n_cols)}{start_row + n_rows}")
    if heat_col and heat_col in df.columns:
        j = list(df.columns).index(heat_col) + 1
        col_letter = get_column_letter(j)
        ws.conditional_formatting.add(
            f"{col_letter}{start_row + 1}:{col_letter}{start_row + n_rows}",
            ColorScaleRule(start_type="min", start_color="FFFFFF",
                           end_type="max", end_color="E8834A"))


def _write_df(writer: pd.ExcelWriter, name: str, df: pd.DataFrame, title: str,
              note: str = "", **style_kw: Any) -> None:
    df.to_excel(writer, sheet_name=name, startrow=2, index=False)
    ws = writer.sheets[name]
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if note:
        ws.cell(row=2, column=1, value=note).font = NOTE_FONT
    _style_sheet(ws, df, start_row=3, **style_kw)


def _stats_frame(kpi: dict[str, Any], keys: list[tuple[str, str]],
                 money: bool = True) -> pd.DataFrame:
    rows = []
    for key, label in keys:
        s = kpi[key]
        rows.append({"Metric": label, "Mean": s["mean"], "Median": s["median"],
                     "Std dev": s["std"], "P5": s["p5"], "P25": s["p25"],
                     "P75": s["p75"], "P95": s["p95"]})
    return pd.DataFrame(rows)


def build_excel_export(
    base_kpi: dict[str, Any],
    baseline: BaselineResult,
    result: SimulationResult,
    summary_text: str,
    scenario_rows: list[dict[str, Any]],
    family_risk: pd.DataFrame,
    binding: pd.DataFrame,
    rankings: dict[str, pd.DataFrame],
    recommendations: list[Recommendation],
    components: pd.DataFrame,
    demand: pd.DataFrame,
    decision: "dict[str, Any] | None" = None,
) -> bytes:
    """Assemble the full workbook and return it as bytes for download.

    The workbook is anchored to the standing base outlook (base, no actions)
    by design — see ARCHITECTURE.md. When an evaluation context is active,
    `decision` adds one explicitly labeled "Decision of Record" sheet
    capturing the context: {"world": str, "actions": [{"name", "cost"}],
    "package_cost": float, "base_kpi": ..., "ctx_kpi": ...,
    "compare": compare_scenarios(base, ctx)}."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # 1. Executive summary
        para_rows = []
        for block in summary_text.split("\n\n"):
            para_rows.append({"Executive summary": block.replace("**", "")})
        _write_df(writer, "Executive Summary", pd.DataFrame(para_rows),
                  "Apex Test Systems — SIOP executive summary",
                  "Generated from simulation results; all data synthetic.")
        ws = writer.sheets["Executive Summary"]
        ws.column_dimensions["A"].width = 110
        for row in ws.iter_rows(min_row=4):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

        # 2. KPI summary
        kpi_rows = [
            ("Q1 revenue (mean)", base_kpi["q1_revenue"]["mean"], MONEY_FMT),
            ("Q1 revenue plan", base_kpi["q1_plan"], MONEY_FMT),
            ("P(Q1 revenue plan)", base_kpi["p_q1_plan"], PCT_FMT),
            ("Q1 revenue at risk (P5 vs plan)", base_kpi["q1_revenue_at_risk"], MONEY_FMT),
            ("FY revenue (mean)", base_kpi["fy_revenue"]["mean"], MONEY_FMT),
            ("FY revenue plan", base_kpi["fy_plan"], MONEY_FMT),
            ("P(FY revenue plan)", base_kpi["p_fy_plan"], PCT_FMT),
            ("FY gross margin (mean)", base_kpi["fy_gm"]["mean"], PCT_FMT),
            ("Gross-margin target", base_kpi["gm_target"], PCT_FMT),
            ("P(gross-margin target)", base_kpi["p_gm_target"], PCT_FMT),
            ("FY operating income (mean)", base_kpi["fy_operating_income"]["mean"], MONEY_FMT),
            ("FY EBITDA proxy (mean)", base_kpi["fy_ebitda"]["mean"], MONEY_FMT),
            ("FY cash-flow proxy (mean)", base_kpi["fy_cash_flow"]["mean"], MONEY_FMT),
            ("Ending inventory (mean)", base_kpi["ending_inventory"]["mean"], MONEY_FMT),
            ("Inventory target", base_kpi["inventory_target"], MONEY_FMT),
            ("P(inventory over target)", base_kpi["p_inventory_over_target"], PCT_FMT),
            ("Inventory turns (mean)", base_kpi["inventory_turns"]["mean"], NUM_FMT),
            ("Working capital (mean)", base_kpi["working_capital"]["mean"], MONEY_FMT),
            ("E&O reserve (mean)", base_kpi["eo_reserve"]["mean"], MONEY_FMT),
            ("Service level (mean fill rate)", base_kpi["service_level"]["mean"], PCT_FMT),
            ("P(missed customer commitment)", base_kpi["p_missed_commitment"], PCT_FMT),
            ("P(component stockout)", base_kpi["p_stockout"], PCT_FMT),
            ("Expected expedite cost (FY)", base_kpi["expedite_cost"]["mean"], MONEY_FMT),
            ("Expected rework cost (FY)", base_kpi["rework_cost"]["mean"], MONEY_FMT),
            ("EMS utilization (mean)", base_kpi["ems_utilization"], PCT_FMT),
            ("Integration utilization (mean)", base_kpi["integration_utilization"], PCT_FMT),
        ]
        kpi_df = pd.DataFrame([{"KPI": k, "Value": v} for k, v, _ in kpi_rows])
        _write_df(writer, "KPI Summary", kpi_df,
                  "KPI summary — base case",
                  f"{base_kpi['n_sims']:,} simulations, seed {result.seed}.")
        ws = writer.sheets["KPI Summary"]
        for i, (_, _, fmt) in enumerate(kpi_rows):
            ws.cell(row=4 + i, column=2).number_format = fmt
        ws.column_dimensions["A"].width = 38

        # 3. Scenario comparison — worlds only: the response-axis columns
        # (always-$0 cost, always-NaN ratio, EV duplicating Δ gross profit at
        # zero cost) are dropped here exactly as they are on-screen
        if scenario_rows:
            # worlds ADD risk: quote the raw delta (positive = more at risk),
            # matching the on-screen worlds-only table; the response-facing
            # strip keeps the sign-flipped "reduced" frame
            sc = pd.DataFrame(scenario_rows).assign(
                d_revenue_at_risk=lambda d: -d["d_revenue_at_risk"]).drop(
                columns=["action_cost", "risk_reduced_per_dollar",
                         "incremental_ev"])
            sc = sc.rename(columns={
                "scenario": "Scenario", "d_q1_revenue": "Δ Q1 revenue",
                "d_fy_revenue": "Δ FY revenue", "d_p_q1_plan": "Δ P(Q1 plan)",
                "d_p_fy_plan": "Δ P(FY plan)", "d_fy_gm": "Δ FY GM",
                "d_gross_profit": "Δ gross profit", "d_inventory": "Δ inventory",
                "d_working_capital": "Δ working capital", "d_expedite": "Δ expedite",
                "d_service": "Δ service level",
                "d_revenue_at_risk": "Δ FY revenue at risk"})
            _write_df(writer, "Scenario Comparison", sc,
                      "Scenario deltas versus base case",
                      money_cols=("Δ Q1 revenue", "Δ FY revenue", "Δ gross profit",
                                  "Δ inventory", "Δ working capital", "Δ expedite",
                                  "Δ FY revenue at risk"),
                      pct_cols=("Δ P(Q1 plan)", "Δ P(FY plan)", "Δ FY GM",
                                "Δ service level"))

        # 4-6. Distribution statistics
        _write_df(writer, "Revenue Distribution",
                  _stats_frame(base_kpi, [("q1_revenue", "Q1 revenue"),
                                          ("q2_revenue", "Q2 revenue"),
                                          ("fy_revenue", "FY revenue")]),
                  "Revenue distribution statistics",
                  money_cols=("Mean", "Median", "Std dev", "P5", "P25", "P75", "P95"))
        gm_df = _stats_frame(base_kpi, [("q1_gm", "Q1 gross margin"),
                                        ("fy_gm", "FY gross margin")])
        _write_df(writer, "Margin Distribution", gm_df,
                  "Gross-margin distribution statistics",
                  pct_cols=("Mean", "Median", "Std dev", "P5", "P25", "P75", "P95"))
        _write_df(writer, "Inventory & WC",
                  _stats_frame(base_kpi, [
                      ("ending_inventory", "Ending inventory (fiscal year-end)"),
                      ("avg_inventory", "Average inventory (FY)"),
                      ("working_capital", "Working capital (fiscal year-end)"),
                      ("eo_reserve", "E&O reserve"),
                      ("expedite_cost", "Expedite cost (FY)"),
                      ("rework_cost", "Rework cost (FY)")]),
                  "Inventory, working-capital and cost statistics",
                  money_cols=("Mean", "Median", "Std dev", "P5", "P25", "P75", "P95"))

        # 7. Monthly baseline plan
        _write_df(writer, "Monthly Baseline", baseline.monthly.round(4),
                  "Deterministic monthly baseline plan",
                  money_cols=tuple(c for c in baseline.monthly.columns if c.endswith("_usd")),
                  pct_cols=("gross_margin", "ems_utilization", "integration_utilization"))

        # 8. Product-family results
        fam = family_risk.copy()
        fam.columns = ["Product family", "Baseline FY revenue", "Expected FY revenue",
                       "P5 FY revenue", "Revenue at risk"]
        _write_df(writer, "Product Families", fam,
                  "Product-family FY results and revenue at risk",
                  money_cols=tuple(fam.columns[1:]), heat_col="Revenue at risk")

        # 9. Customer-group demand
        cust = (demand.assign(units=demand["base_forecast_units"] + demand["backlog_units"],
                              revenue=lambda d: d["units"] * d["asp_usd"])
                .groupby(["customer", "customer_group", "region"], as_index=False)
                .agg(fy_units=("units", "sum"), fy_revenue=("revenue", "sum"),
                     avg_push_prob=("push_out_prob", "mean"),
                     avg_cancel_prob=("cancel_prob", "mean")))
        cust.columns = ["Customer", "Group", "Region", "Horizon units",
                        "Horizon revenue", "Avg push-out prob", "Avg cancel prob"]
        _write_df(writer, "Customer Groups", cust,
                  "Customer demand overview (18-month horizon)",
                  money_cols=("Horizon revenue",),
                  pct_cols=("Avg push-out prob", "Avg cancel prob"))

        # 10. EMS capacity
        site = baseline.site_load.T.copy()
        site.insert(0, "Month", month_labels())
        cap_t = baseline.site_capacity.T
        for s in baseline.site_capacity.index:
            site[f"{s} capacity"] = cap_t[s].values
        _write_df(writer, "EMS Capacity", site.round(1),
                  "EMS site load vs effective capacity (std-equivalent units)")

        # 11. Component risk
        comp = binding.merge(
            components[["component", "supplier", "supplier_region", "lead_time_weeks",
                        "allocation_risk", "disruption_prob_monthly",
                        "monthly_requirement_units"]], on="component", how="left")
        comp.columns = ["Component", "Binding frequency", "Supplier", "Region",
                        "Lead time (wk)", "Allocation risk", "Disruption prob",
                        "Monthly req (units)"]
        _write_df(writer, "Component Risk", comp,
                  "Critical-component risk (simulation binding frequency)",
                  pct_cols=("Binding frequency", "Allocation risk", "Disruption prob"),
                  heat_col="Binding frequency")

        # 12. Sensitivity rankings
        sens_rows = []
        for outcome, df in rankings.items():
            for _, r in df.head(8).iterrows():
                sens_rows.append({"Outcome": outcome, "Driver": r["driver"],
                                  "Spearman rho": r["spearman_rho"]})
        _write_df(writer, "Sensitivity", pd.DataFrame(sens_rows),
                  "Driver rankings (Spearman rank correlation)",
                  "Association, not causation: drivers share common factors by design.")

        # 13. Recommendations
        if recommendations:
            rec_df = pd.DataFrame([{
                "Rank": i + 1, "Recommendation": r.title, "Detected risk": r.risk,
                "Trigger threshold": r.threshold,
                "Expected value": r.expected_value_usd,
                "Revenue protected": r.revenue_protected_usd,
                "Gross profit protected": r.gross_profit_protected_usd,
                "Δ P(plan)": r.prob_plan_improvement,
                "Incremental cost": r.incremental_cost_usd,
                "Δ inventory": r.inventory_change_usd,
                "Δ working capital": r.working_capital_change_usd,
                "Confidence": r.confidence, "Caveat": r.caveat,
            } for i, r in enumerate(recommendations)])
            _write_df(writer, "Recommendations", rec_df,
                      "Ranked management recommendations (modeled impacts)",
                      money_cols=("Expected value", "Revenue protected",
                                  "Gross profit protected", "Incremental cost",
                                  "Δ inventory", "Δ working capital"),
                      pct_cols=("Δ P(plan)",))

        # 13b. Decision of record — only when an evaluation context is active.
        # The rest of the workbook stays anchored to (base, ∅); this sheet is
        # the explicitly labeled record of the decision the meeting evaluated.
        if decision is not None:
            b_kpi, c_kpi = decision["base_kpi"], decision["ctx_kpi"]
            pkg = decision["actions"]
            head_rows = [
                {"Item": "World evaluated (scenario)", "Value": decision["world"]},
                {"Item": "Response package",
                 "Value": f"{len(pkg)} action(s)" if pkg else "none"},
            ]
            head_rows += [
                {"Item": f"  package action {i}",
                 "Value": f"{a['name']} — cost ${a['cost'] / 1e6:,.1f}M"}
                for i, a in enumerate(pkg, 1)]
            head_rows.append(
                {"Item": "Package decision cost (charged to Q1 opex)",
                 "Value": f"${decision['package_cost'] / 1e6:,.1f}M"})
            head = pd.DataFrame(head_rows)
            _write_df(writer, "Decision of Record", head,
                      "Decision of record — evaluated context (world, response)",
                      "Every other sheet stays anchored to the standing base "
                      "outlook (no scenario, no actions) by design; this sheet "
                      "records the context evaluated at export time.")
            ws = writer.sheets["Decision of Record"]
            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 52

            kpi_keys = [
                ("Expected Q1 revenue", lambda k: k["q1_revenue"]["mean"], MONEY_FMT),
                ("P(Q1 revenue plan)", lambda k: k["p_q1_plan"], PCT_FMT),
                ("Expected FY revenue", lambda k: k["fy_revenue"]["mean"], MONEY_FMT),
                ("P(FY revenue plan)", lambda k: k["p_fy_plan"], PCT_FMT),
                ("Expected FY gross margin", lambda k: k["fy_gm"]["mean"], PCT_FMT),
                ("FY revenue at risk (P5 vs plan)",
                 lambda k: k["fy_revenue_at_risk"], MONEY_FMT),
                ("Expected FY gross profit",
                 lambda k: k["fy_gross_profit"]["mean"], MONEY_FMT),
                ("Year-end inventory (mean)",
                 lambda k: k["ending_inventory"]["mean"], MONEY_FMT),
                ("FY service level (fill rate)",
                 lambda k: k["service_level"]["mean"], PCT_FMT),
            ]
            cmp_rows = [{"KPI": label,
                         "Base outlook (base, no actions)": f(b_kpi),
                         "Conditioned on the decision": f(c_kpi),
                         "Δ": f(c_kpi) - f(b_kpi)}
                        for label, f, _ in kpi_keys]
            fmts = [fmt for _, _, fmt in kpi_keys]
            if pkg:
                cmp_rows.append({
                    "KPI": "FY incremental EV (net of package cost)",
                    "Base outlook (base, no actions)": None,
                    "Conditioned on the decision": None,
                    "Δ": decision["compare"]["incremental_ev"]})
                fmts.append(MONEY_FMT)
            cmp_df = pd.DataFrame(cmp_rows)
            title_row = len(head) + 6          # two blank rows below table 1
            hdr = title_row + 1
            cmp_df.to_excel(writer, sheet_name="Decision of Record",
                            startrow=hdr - 1, index=False)
            ws.cell(row=title_row, column=1,
                    value="Conditioned outlook vs the standing base outlook"
                    ).font = TITLE_FONT
            for j in range(1, len(cmp_df.columns) + 1):
                cell = ws.cell(row=hdr, column=j)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if j > 1:
                    ws.column_dimensions[get_column_letter(j)].width = 28
            for i, fmt in enumerate(fmts):
                for j in range(1, len(cmp_df.columns) + 1):
                    cell = ws.cell(row=hdr + 1 + i, column=j)
                    cell.border = BORDER
                    if j > 1 and isinstance(cell.value, (int, float)):
                        cell.number_format = fmt

        # 14. Assumptions
        prod_cols = components.columns.tolist()
        _write_df(writer, "Assumptions - Components", components,
                  "Component master assumptions (synthetic)")
        # 15. Methodology
        method = pd.DataFrame({"Methodology notes": [
            "All data is synthetic; no real company data is used.",
            "Baseline: deterministic greedy allocation (backlog first, customer "
            "priority, requested date, contribution margin), respecting component, "
            "EMS and integration constraints.",
            "Monte Carlo: correlated common-factor model (8 factors, AR(1) "
            "persistence); bounded distributions (lognormal multipliers, Bernoulli "
            "disruptions, beta-shaped slip fractions).",
            "Hybrid granularity: simulation runs at family x month with all 30 "
            "critical components and site-level capacity; within-month rationing is "
            "proportional (documented approximation of baseline priority order).",
            "Financials: revenue = recognized units x realized ASP; COGS includes "
            "material (with PPV/FX), conversion, integration, freight, warranty, "
            "scrap, rework, expedite and overtime premiums.",
            "Working capital = inventory + simplified receivables (DSO) - simplified "
            "payables (DPO). Cash proxy = EBITDA - ΔWC - capex - cash taxes.",
            "E&O: reserve rate applied to critical-component stock above 2.5 months "
            "of forward usage at fiscal year end, plus 5% of aged finished goods.",
            "Sensitivity: Spearman rank correlation — association, not causation.",
            "Limitations: monthly buckets, family-level simulation, no MILP "
            "optimization, simplified revenue recognition, no FX balance-sheet "
            "effects.",
        ]})
        _write_df(writer, "Methodology", method, "Methodology and simplifications")
        writer.sheets["Methodology"].column_dimensions["A"].width = 110
        for row in writer.sheets["Methodology"].iter_rows(min_row=4):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    return buf.getvalue()
