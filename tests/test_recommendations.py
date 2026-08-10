"""Recommendations must cite actual modeled outcomes; exports must build."""
from __future__ import annotations

import numpy as np
import pytest
from openpyxl import load_workbook

from src.executive_report import ReportContext, get_provider
from src.exports import build_excel_export
from src.recommendations import build_recommendations, detect_risks
from src.scenarios import kpi_summary, management_actions
from src.sensitivity import (
    all_driver_rankings,
    binding_components,
    family_revenue_at_risk,
    monthly_capacity_risk,
)
from src.simulation import run_simulation


@pytest.fixture(scope="module")
def pipeline(data, config, baseline, base_result):
    base_kpi = kpi_summary(base_result, baseline, config)
    binding = binding_components(base_result)
    action_results = {}
    for name, spec in list(management_actions().items())[:5]:
        r = run_simulation(data, config, baseline, params=spec.overrides,
                           n_sims=base_result.n_sims, seed=base_result.seed,
                           scenario_name=name)
        action_results[name] = (kpi_summary(r, baseline, config), spec)
    recs = build_recommendations(base_kpi, action_results, binding)
    return base_kpi, binding, action_results, recs


def test_recommendations_reference_modeled_outcomes(pipeline):
    base_kpi, _, action_results, recs = pipeline
    by_name = {name: kpi for name, (kpi, _) in action_results.items()}
    for rec in recs:
        assert rec.title in by_name, "recommendation must map to a simulated action"
        kpi, spec = action_results[rec.title]
        expected_ev = (kpi["fy_gross_profit"]["mean"]
                       - base_kpi["fy_gross_profit"]["mean"]
                       - spec.action_cost_usd)
        assert rec.expected_value_usd == pytest.approx(expected_ev)
        assert rec.incremental_cost_usd == spec.action_cost_usd


def test_recommendations_materiality_gates(pipeline):
    _, _, _, recs = pipeline
    for rec in recs:
        assert rec.expected_value_usd > -1.0e6
        assert (rec.prob_plan_improvement >= 0.01
                or rec.expected_value_usd >= 2.0e6)


def test_risk_detection_thresholds(pipeline):
    base_kpi, binding, _, _ = pipeline
    risks = detect_risks(base_kpi, binding)
    for r in risks:
        assert r["threshold"]          # every risk carries its trigger
        assert r["risk"]


def test_executive_summary_is_dynamic(data, config, baseline, base_result, pipeline):
    """Summary must change when results change (different seed)."""
    base_kpi, binding, _, recs = pipeline
    provider = get_provider("rules")

    def make_summary(kpi, result):
        ctx = ReportContext(
            kpi=kpi, risks=detect_risks(kpi, binding), binding=binding,
            family_risk=family_revenue_at_risk(result, baseline),
            capacity_risk=monthly_capacity_risk(result), recommendations=recs)
        return provider.executive_summary(ctx)

    s1 = make_summary(base_kpi, base_result)
    other = run_simulation(data, config, baseline, n_sims=800, seed=1234)
    s2 = make_summary(kpi_summary(other, baseline, config), other)
    assert s1 != s2
    word_count = len(s1.split())
    assert 200 <= word_count <= 550


def test_executive_summary_conditioned_on_scenario(baseline, base_result, pipeline):
    """A conditioned summary must name the scenario and flag base-anchored
    decisions; the unconditioned summary must do neither."""
    base_kpi, binding, _, recs = pipeline
    provider = get_provider("rules")

    def make_summary(conditioned_on):
        ctx = ReportContext(
            kpi=base_kpi, risks=detect_risks(base_kpi, binding), binding=binding,
            family_risk=family_revenue_at_risk(base_result, baseline),
            capacity_risk=monthly_capacity_risk(base_result),
            recommendations=recs, conditioned_on=conditioned_on)
        return provider.executive_summary(ctx)

    plain = make_summary(None)
    conditioned = make_summary("Critical FPGA Shortage")
    assert "conditioned on Critical FPGA Shortage" in conditioned
    assert "standing base-case outlook" in conditioned
    assert "conditioned on" not in plain
    assert "standing base-case outlook" not in plain


def test_excel_export_builds(tmp_path, data, config, baseline, base_result, pipeline):
    base_kpi, binding, _, recs = pipeline
    xl = build_excel_export(
        base_kpi, baseline, base_result, "Summary paragraph.\n\nSecond paragraph.",
        [], family_revenue_at_risk(base_result, baseline), binding,
        all_driver_rankings(base_result), recs, data.components, data.demand)
    path = tmp_path / "export.xlsx"
    path.write_bytes(xl)
    wb = load_workbook(path)
    for sheet in ["Executive Summary", "KPI Summary", "Revenue Distribution",
                  "Monthly Baseline", "Product Families", "Component Risk",
                  "Sensitivity", "Recommendations", "Methodology"]:
        assert sheet in wb.sheetnames
